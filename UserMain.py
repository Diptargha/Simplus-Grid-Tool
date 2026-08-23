"""UserMain.py — single entry point for the Python SimplusGT pipeline.

Edit the USER SETTINGS block below, then run this file (PyCharm / VS Code /
``python UserMain.py``). No CLI is required.

Mirrors MATLAB ``UserMain.m``: pick a case, toggle which analyses and plots
to run.
"""

from __future__ import annotations

import json
from pathlib import Path

import matplotlib

# ---------------------------------------------------------------------------
# USER SETTINGS — customise these
# ---------------------------------------------------------------------------

# Case name (short name, as in MATLAB UserDataName) or a full/relative path.
# Short names are resolved under Examples/** and the repo root.
#
# Ac:
#   "SgInfiniteBus", "GflInverterInfiniteBus", "GfmInverterInfiniteBus",
#   "BessInfiniteBus", "PV_GfmInfiniteBus", "PV_GflInfiniteBus",
#   "WtGfmInfiniteBus", "WtGflInfiniteBus",
#   "IEEE_14Bus", "IEEE_30Bus", "IEEE_57Bus", "AU14Gen_59Bus", "NETS_NYPS_68Bus",
#   "BESS_Plant_10Bus", "PV_Plant_10Bus", "PV_BESS_Hybrid_Plant_10Bus",
#   "WSCC_9Bus",
# Dc:
#   "GfdBuckInfiniteBus", "TwoBusGfdBuck",
# Hybrid:
#   "Hybrid_4Bus", "HVDC_Infbus_4Bus", "HVDC_SG_4Bus", "MTDC_Infbus_4Bus",
# Default 4-bus:
#   "UserData"
USER_DATA_NAME = "IEEE_14Bus"

# Preferred file type when resolving a short name: "json" or "excel" (xlsx/xlsm).
USER_DATA_TYPE = "json"

# --- Core pipeline (always runs) -------------------------------------------
# Power flow + apparatus/network DSS + eigenvalues.

# --- Fundamental plots (MATLAB Main.m) -------------------------------------
ENABLE_PLOT_POLE = True
ENABLE_PLOT_ADMITTANCE = False          # Ydd + complex-vector Ydq+
ENABLE_PLOT_ADMITTANCE_DQ_AXES = False  # dd / dq / qd / qq (Modal BodeDraw)
ENABLE_PLOT_GRID_STRENGTH = True

# --- Greybox / modal analysis ----------------------------------------------
ENABLE_GREYBOX = True
ENABLE_GREYBOX_APP_LAYER1 = True   # apparatus Layer 1
ENABLE_GREYBOX_APP_LAYER2 = True   # apparatus Layer 2
ENABLE_GREYBOX_APP_LAYER3 = True  # apparatus Layer 3 (also adds Layer3 sheet to Excel export)
ENABLE_GREYBOX_SENS_LAYER12 = True   # sensitivity Layer 1/2
ENABLE_GREYBOX_SENS_LAYER3 = True   # sensitivity Layer 3
# Mode selection (0-based indices into state-space eig(A), same order as greybox):
#   "auto"  — top GREYBOX_MODE_MAX_COUNT least-damped oscillatory modes
#   "freq"  — top N least-damped oscillatory modes in GREYBOX_MODE_FREQ_* band (fallback: "auto")
#   (157, 159) or "157,159" — explicit raw indices
# Band also applies when GREYBOX_MODE_FREQ_MIN_HZ and GREYBOX_MODE_FREQ_MAX_HZ are both set
# (even if GREYBOX_MODES is "auto").
GREYBOX_MODES = "freq"
GREYBOX_MODE_FREQ_MIN_HZ = 1  # e.g. 1.0 to select modes by frequency band
GREYBOX_MODE_FREQ_MAX_HZ = 10  # e.g. 10.0
GREYBOX_MODE_MAX_COUNT = 3
# Ysys/Zsys Bode frequency grid (unrelated to mode selection):
GREYBOX_FREQ_MIN_HZ = 1
GREYBOX_FREQ_MAX_HZ = 1000
# Linear grid step (Hz). Overrides log-spaced GREYBOX_FREQ_COUNT when set.
GREYBOX_FREQ_SPACING_HZ = 1.0
GREYBOX_FREQ_COUNT = 1000  # used only if GREYBOX_FREQ_SPACING_HZ is None

# --- Greybox plots ---------------------------------------------------------
ENABLE_PLOT_GREYBOX = True  # Ysys/Zsys Bode + Layer 1/2 charts

# --- Interactive HTML dashboard (Plotly) -----------------------------------
# Single browser file with pole / strength / greybox Bode / Layer 1–2 charts.
# Prefer this over matplotlib for Layer 1/2 (labels stay readable via hover).
ENABLE_HTML_DASHBOARD = True
HTML_DASHBOARD_DIR = "Results"  # writes <case>_dashboard.html

# --- Exports ---------------------------------------------------------------
ENABLE_EXPORT_DASHBOARD = False       # JSON dashboard
ENABLE_EXPORT_GREYBOX_JSON = False    # JSON greybox (includes Zsys)
ENABLE_EXPORT_GREYBOX_EXCEL = True    # Excel greybox (Ysys/Zsys, Eigenvalues, StatePF, Layers when enabled)
EXPORT_DIR = "Results"

# --- Display / save --------------------------------------------------------
SHOW_PLOTS = True          # open HTML dashboard (and matplotlib only if HTML is off)
SAVE_PLOTS = False         # also write matplotlib PNGs under PLOT_DIR
PLOT_DIR = "Results/plots"

# Non-interactive backend when not showing matplotlib windows.
if not SHOW_PLOTS or ENABLE_HTML_DASHBOARD:
    matplotlib.use("Agg")

# ---------------------------------------------------------------------------
# Implementation — usually no need to edit below
# ---------------------------------------------------------------------------

from simplusgt.analysis import stability_report
from simplusgt.export import export_greybox_excel, greybox_result_to_dashboard_data, result_to_dashboard_data
from simplusgt.greybox import FrequencyGrid, GreyboxConfig, GreyboxLayerSelection, run_greybox
from simplusgt.html_dashboard import write_analysis_dashboard
from simplusgt.pipeline import run_case
from simplusgt.plotting import plot_case_fundamentals, plot_greybox_summary

ROOT = Path(__file__).resolve().parent


def resolve_case_path(name: str, preferred: str = "json") -> Path:
    """Resolve a short case name or path to an existing case file."""

    candidate = Path(name)
    if candidate.is_file():
        return candidate.resolve()
    rooted = ROOT / name
    if rooted.is_file():
        return rooted.resolve()

    stem = name
    for suffix in (".json", ".xlsx", ".xlsm", ".xls"):
        if stem.lower().endswith(suffix):
            stem = stem[: -len(suffix)]
            break

    preferred = preferred.lower().strip()
    if preferred in {"excel", "xlsx", "xlsm", "1"}:
        order = (".xlsx", ".xlsm", ".xls", ".json")
    else:
        order = (".json", ".xlsx", ".xlsm", ".xls")

    search_roots = [ROOT, ROOT / "Examples"]
    matches: list[Path] = []
    for ext in order:
        pattern = f"**/{stem}{ext}"
        for base in search_roots:
            matches.extend(sorted(base.glob(pattern)))
        if matches:
            break
    # Also allow bare name in repo root (UserData.json).
    for ext in order:
        direct = ROOT / f"{stem}{ext}"
        if direct.is_file():
            return direct.resolve()

    if not matches:
        raise FileNotFoundError(
            f'Could not find case "{name}" under {ROOT} or Examples/. '
            "Use a short name (e.g. IEEE_14Bus) or a full path to a .json/.xlsx file."
        )
    return matches[0].resolve()


def _system_eigenvalues(run_result):
    """Eigenvalues in the same order as greybox modal analysis (``np.linalg.eig(A)``)."""

    import numpy as np

    model = run_result.whole_system_dss
    try:
        matrix_a, _, _, _ = model.to_state_space()
        values, _ = np.linalg.eig(matrix_a)
    except np.linalg.LinAlgError:
        from scipy.linalg import eig

        values, _, _ = eig(model.A, model.E, left=False, right=False)
    return np.asarray(values, dtype=complex).ravel()


def _oscillatory_modes_scored(eigenvalues) -> list[tuple[int, float, float]]:
    import numpy as np

    scored: list[tuple[int, float, float]] = []
    for idx, value in enumerate(np.asarray(eigenvalues, dtype=complex).ravel()):
        if not np.isfinite(value):
            continue
        omega = float(np.imag(value))
        if omega <= 1e-6:
            continue
        sigma = float(np.real(value))
        mag = abs(complex(value))
        freq_hz = abs(omega / (2 * np.pi))
        zeta = (-sigma / mag) if mag > 0 else float("inf")
        scored.append((idx, zeta, freq_hz))
    return scored


def _auto_least_damped_modes(eigenvalues, *, max_count: int = 3) -> tuple[int, ...]:
    import numpy as np

    scored = _oscillatory_modes_scored(eigenvalues)
    scored.sort(key=lambda item: item[1])
    if scored:
        return tuple(idx for idx, _, _ in scored[: max(1, int(max_count))])
    values = np.asarray(eigenvalues, dtype=complex).ravel()
    finite = [idx for idx, value in enumerate(values) if np.isfinite(value)]
    return tuple(finite[: max(1, int(max_count))]) if finite else (0,)


def _modes_in_frequency_band(
    eigenvalues,
    freq_min_hz: float,
    freq_max_hz: float,
    *,
    max_count: int = 3,
) -> tuple[int, ...] | None:
    if freq_min_hz > freq_max_hz:
        freq_min_hz, freq_max_hz = freq_max_hz, freq_min_hz
    in_band = [
        (idx, zeta)
        for idx, zeta, freq_hz in _oscillatory_modes_scored(eigenvalues)
        if freq_min_hz <= freq_hz <= freq_max_hz
    ]
    if not in_band:
        return None
    in_band.sort(key=lambda item: item[1])
    return tuple(idx for idx, _ in in_band[: max(1, int(max_count))])


def parse_modes(
    spec: str | tuple[int, ...] | list[int],
    run_result,
    *,
    freq_min_hz: float | None = None,
    freq_max_hz: float | None = None,
    max_count: int = 3,
) -> tuple[int, ...]:
    import numpy as np

    eigenvalues = _system_eigenvalues(run_result)
    text = str(spec).strip().lower() if not isinstance(spec, (tuple, list)) else ""
    use_band = text == "freq" or (freq_min_hz is not None and freq_max_hz is not None)
    if use_band:
        fmin = float(freq_min_hz if freq_min_hz is not None else 0.0)
        fmax = float(freq_max_hz if freq_max_hz is not None else float("inf"))
        band_modes = _modes_in_frequency_band(eigenvalues, fmin, fmax, max_count=max_count)
        if band_modes:
            return band_modes
        print(
            f"  No oscillatory modes in {fmin:g}–{fmax:g} Hz; "
            f"falling back to top {max_count} least-damped modes."
        )
        return _auto_least_damped_modes(eigenvalues, max_count=max_count)

    if isinstance(spec, (tuple, list)):
        return tuple(int(v) for v in spec)
    if text == "auto":
        return _auto_least_damped_modes(eigenvalues, max_count=max_count)
    return tuple(int(part.strip()) for part in text.split(",") if part.strip())


def main() -> None:
    case_path = resolve_case_path(USER_DATA_NAME, USER_DATA_TYPE)
    print("=" * 50)
    print("SimplusGT (Python)")
    print("=" * 50)
    print(f"Case: {case_path}")

    result = run_case(case_path)
    report = stability_report(result.eigenvalues)
    print(f"Buses: {result.netlists.buses.shape[0]}")
    print(f"Lines: {result.netlists.lines.shape[0]}")
    print(f"States: {result.whole_system_dss.nx}")
    print(f"Stable: {report.stable}")
    if result.warnings:
        print("Warnings:")
        for warning in result.warnings:
            print(f"  - {warning}")

    plot_dir = ROOT / PLOT_DIR if SAVE_PLOTS else None
    # When the HTML dashboard is enabled, matplotlib is PNG-only (no SciView spam).
    show_mpl = bool(SHOW_PLOTS and not ENABLE_HTML_DASHBOARD)
    if any(
        (
            ENABLE_PLOT_POLE,
            ENABLE_PLOT_ADMITTANCE,
            ENABLE_PLOT_ADMITTANCE_DQ_AXES,
            ENABLE_PLOT_GRID_STRENGTH,
        )
    ) and (SAVE_PLOTS or show_mpl):
        print("\nFundamental plots...")
        saved = plot_case_fundamentals(
            result,
            output_dir=plot_dir,
            show=show_mpl,
            include_pole=ENABLE_PLOT_POLE,
            include_admittance=ENABLE_PLOT_ADMITTANCE,
            include_dq_axes=ENABLE_PLOT_ADMITTANCE_DQ_AXES,
            include_strength=ENABLE_PLOT_GRID_STRENGTH,
        )
        for key, path in saved.items():
            if path is not None:
                print(f"  {key}: {path}")

    greybox = None
    need_greybox = (
        ENABLE_GREYBOX
        or ENABLE_PLOT_GREYBOX
        or ENABLE_EXPORT_GREYBOX_JSON
        or ENABLE_EXPORT_GREYBOX_EXCEL
        or (ENABLE_HTML_DASHBOARD and ENABLE_PLOT_GREYBOX)
    )
    if need_greybox:
        print("\nGreybox analysis...")
        modes = parse_modes(
            GREYBOX_MODES,
            result,
            freq_min_hz=GREYBOX_MODE_FREQ_MIN_HZ,
            freq_max_hz=GREYBOX_MODE_FREQ_MAX_HZ,
            max_count=GREYBOX_MODE_MAX_COUNT,
        )
        freq_grid = FrequencyGrid(
            min_hz=GREYBOX_FREQ_MIN_HZ,
            max_hz=GREYBOX_FREQ_MAX_HZ,
            count=GREYBOX_FREQ_COUNT,
            spacing_hz=GREYBOX_FREQ_SPACING_HZ,
        )
        n_freq = int(freq_grid.frequencies().size)
        greybox = run_greybox(
            GreyboxConfig(
                case_path=case_path,
                layers=GreyboxLayerSelection(
                    apparatus_layer1=ENABLE_GREYBOX_APP_LAYER1,
                    apparatus_layer2=ENABLE_GREYBOX_APP_LAYER2,
                    apparatus_layer3=ENABLE_GREYBOX_APP_LAYER3,
                    sensitivity_layer12=ENABLE_GREYBOX_SENS_LAYER12,
                    sensitivity_layer3=ENABLE_GREYBOX_SENS_LAYER3,
                ),
                modes=modes,
                frequency_grid=freq_grid,
                # Keep full grid (do not downsample Ysys/Zsys to the default 80 points).
                max_admittance_samples=max(n_freq, 80),
            )
        )
        print(f"  Modes: {modes}")
        print(f"  Apparatus Layer-1 modes: {len(greybox.modes)}")
        print(f"  Sensitivity results: {len(greybox.sensitivity)}")
        if greybox.warnings:
            for warning in greybox.warnings:
                if warning not in result.warnings:
                    print(f"  - {warning}")

        # Matplotlib greybox PNGs only when saving; interactive Layer charts use HTML.
        if ENABLE_PLOT_GREYBOX and SAVE_PLOTS:
            print("Greybox plots (PNG)...")
            saved_gb = plot_greybox_summary(greybox, output_dir=plot_dir, show=False)
            for key, path in saved_gb.items():
                if path is not None:
                    print(f"  {key}: {path}")
        elif ENABLE_PLOT_GREYBOX and show_mpl:
            print("Greybox plots...")
            saved_gb = plot_greybox_summary(greybox, output_dir=None, show=True)
            for key, path in saved_gb.items():
                if path is not None:
                    print(f"  {key}: {path}")

    if ENABLE_HTML_DASHBOARD:
        html_dir = ROOT / HTML_DASHBOARD_DIR
        html_path = html_dir / f"{case_path.stem}_dashboard.html"
        print("\nHTML dashboard...")
        written = write_analysis_dashboard(
            result,
            greybox,
            output_path=html_path,
            include_pole=ENABLE_PLOT_POLE,
            include_strength=ENABLE_PLOT_GRID_STRENGTH,
            include_admittance=ENABLE_PLOT_ADMITTANCE,
            include_dq_axes=ENABLE_PLOT_ADMITTANCE_DQ_AXES,
            include_greybox=ENABLE_PLOT_GREYBOX,
            case_label=case_path.stem,
            open_browser=SHOW_PLOTS,
        )
        print(f"  {written}")

    export_dir = ROOT / EXPORT_DIR
    if ENABLE_EXPORT_DASHBOARD:
        export_dir.mkdir(parents=True, exist_ok=True)
        out = export_dir / f"{case_path.stem}_dashboard.json"
        payload = result_to_dashboard_data(result, case_path=str(case_path))
        out.write_text(json.dumps(payload, indent=2, allow_nan=False), encoding="utf-8")
        print(f"\nDashboard export: {out}")

    if ENABLE_EXPORT_GREYBOX_JSON:
        if greybox is None:
            raise RuntimeError("ENABLE_EXPORT_GREYBOX_JSON requires greybox analysis")
        export_dir.mkdir(parents=True, exist_ok=True)
        out = export_dir / f"{case_path.stem}_greybox.json"
        payload = greybox_result_to_dashboard_data(greybox)
        out.write_text(json.dumps(payload, indent=2, allow_nan=False), encoding="utf-8")
        print(f"Greybox JSON export: {out}")

    if ENABLE_EXPORT_GREYBOX_EXCEL:
        if greybox is None:
            raise RuntimeError("ENABLE_EXPORT_GREYBOX_EXCEL requires greybox analysis")
        export_dir.mkdir(parents=True, exist_ok=True)
        out = export_dir / f"{case_path.stem}_greybox.xlsx"
        export_greybox_excel(greybox, out)
        print(f"Greybox Excel export: {out}")
        # Surface Layer sheet presence in the console (tabs can be easy to miss in Excel).
        try:
            import openpyxl

            wb = openpyxl.load_workbook(out, read_only=True)
            layer_tabs = [n for n in wb.sheetnames if n.startswith("Layer") or n.startswith("Sens_")]
            print(f"  Sheets: {', '.join(wb.sheetnames)}")
            print(f"  Layer tabs: {', '.join(layer_tabs) if layer_tabs else '(none — enable ENABLE_GREYBOX_*_LAYER*)'}")
            wb.close()
        except Exception as exc:  # noqa: BLE001
            print(f"  (could not list sheets: {exc})")

    print("\nDone.")


if __name__ == "__main__":
    main()
